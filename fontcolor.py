import openpyxl

# Specify the input Excel file path
input_file = r"C:\Users\Sirsha\Desktop\LAB WORK NISER\Sohini\20062023\time_intervals\common_reg.xlsx"

# Load the Excel file
workbook = openpyxl.load_workbook(input_file)

# Iterate over each sheet in the workbook
for sheet in workbook.sheetnames:
    # Get the current sheet
    current_sheet = workbook[sheet]
    
    # Iterate over each row in the sheet
    for row in current_sheet.iter_rows():
        # Iterate over each cell in the row
        for cell in row:
            # Check if the cell value is "UP"
            if cell.value == 'UP':
                # Set font color to green
                cell.font = openpyxl.styles.Font(color="00AA00")  # Green color code
            # Check if the cell value is "DOWN"
            elif cell.value == 'DOWN':
                # Set font color to red
                cell.font = openpyxl.styles.Font(color="FF0000")  # Red color code

# Save the changes to a new Excel file
output_file = r"C:\Users\Sirsha\Desktop\LAB WORK NISER\Sohini\20062023\time_intervals\common_regulation.xlsx"
workbook.save(output_file)
